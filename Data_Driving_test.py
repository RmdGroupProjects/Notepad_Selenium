import openpyxl
path=r"C:\Users\ram\Desktop\OrangeHrm.xlsx"
work_book=openpyxl.load_workbook(path)
sheet=work_book.active
row=sheet.max_row
col=sheet.max_column
print(row)
print(col)
for i in range(1,row+1):
    for j in range(1,col+1):
        print(sheet.cell(row=i,column=j).value,end="      ")
    print()